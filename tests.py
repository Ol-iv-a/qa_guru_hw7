import csv
from io import TextIOWrapper
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook
from config import ZIP_DIR


def test_pdf_content(create_archive):
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open("file.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()

            assert "Тестовый PDF" in text

def test_xlsx_content(create_archive):
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open("file.xlsx") as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active

            assert sheet.max_row == 8
            assert sheet.cell(row=7, column=1).value == 300987
            assert sheet.cell(row=7, column=2).value == "Смоленск 2"

def test_csv_content(create_archive):
    with (ZipFile(ZIP_DIR) as zip_file):
        with zip_file.open("file.csv") as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            rows = list(csvreader)

            assert len(rows) == 8
            assert rows[6] == ['300987', 'Смоленск 2']
